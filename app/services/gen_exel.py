"""Модуль генерации Excel-файлов на основе записей BOM из базы данных."""

from datetime import datetime
from tempfile import NamedTemporaryFile

import openpyxl
from fastapi import HTTPException
from fastapi.responses import FileResponse
from openpyxl.styles import Font, Alignment
from sqlmodel import select

from app.models import DemoRecord, Upload

CM_TO_POINTS = 28.35  # 1 см ≈ 28.35 пунктов (points)
ROW_HEIGHT_CM = 0.8
ROW_HEIGHT_PT = ROW_HEIGHT_CM * CM_TO_POINTS


def split_item(item):
    """Разделяет префикс и числовую часть: R12 -> ('R', 12)."""
    if not item:
        return None, None

    prefix = ''
    num_str = ''
    for char in item:
        if char.isalpha():
            prefix += char
        elif char.isdigit():
            num_str += char

    num = int(num_str) if num_str else None
    return prefix, num


def shorten_ranges(s):
    """
    Сокращает список компонентов с одинаковым префиксом в диапазоны.
    Пример:
        "R1, R2, R3, R5" -> "R1-R3, R5"
    """
    if not s:
        return ""

    items = [item.strip() for item in s.split(",")]
    if not items:
        return ""

    has_numbers = any(split_item(item)[1] is not None for item in items)
    if not has_numbers:
        return s

    valid_items = []
    for item in items:
        prefix, num = split_item(item)
        if num is not None:
            valid_items.append((prefix, num))

    if not valid_items:
        return ""

    prefix = valid_items[0][0]
    numbers = [num for _, num in valid_items]
    numbers.sort()

    ranges = []
    start = numbers[0]
    prev = start

    for num in numbers[1:]:
        if num == prev + 1:
            prev = num
        else:
            if start == prev:
                ranges.append(f"{prefix}{start}")
            else:
                ranges.append(f"{prefix}{start}-{prefix}{prev}")
            start = num
            prev = num

    if start == prev:
        ranges.append(f"{prefix}{start}")
    else:
        ranges.append(f"{prefix}{start}-{prefix}{prev}")

    return ", ".join(ranges)


def generate_excel_for_upload(upload_id: int, session):
    """
    Генерирует Excel файл для указанного upload_id и возвращает FileResponse.
    """
    try:
        records = session.exec(
            select(DemoRecord).where(DemoRecord.upload_id == upload_id)
        ).all()

        if not records:
            raise HTTPException(status_code=404, detail="Записей нет")

        upload = session.get(Upload, upload_id)
        if not upload:
            raise HTTPException(status_code=404, detail="Загрузка не найдена")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM Data"

        default_font = Font(name='GOST Type A', size=12, italic=True)
        align_center = Alignment(horizontal="center", vertical="center")

        headers = ["Поз. обозначение", "Наименование", "Кол.", "Примечание"]
        ws.append(headers)
        ws.row_dimensions[1].height = ROW_HEIGHT_PT  # высота строки заголовка

        for cell in ws[1]:
            cell.font = default_font
            cell.alignment = align_center

        column_widths = {
            'A': 2.0 * CM_TO_POINTS,
            'B': 11.0 * CM_TO_POINTS,
            'C': 1.0 * CM_TO_POINTS,
            'D': 4.5 * CM_TO_POINTS
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width / 7

        for record in records:
            row = [
                shorten_ranges(record.designator),
                f"{record.ad_class} {record.ad_bom} {record.ad_ss}".strip(),
                record.quantity,
                record.ad_note
            ]
            ws.append(row)

            current_row = ws.max_row
            ws.row_dimensions[current_row].height = ROW_HEIGHT_PT  # высота строки

            for col_idx, cell in enumerate(ws[current_row], start=1):
                cell.font = default_font
                if col_idx in (1, 3):
                    cell.alignment = align_center

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_file_path = temp_file.name
            wb.save(temp_file_path)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"bom_upload_{upload_id}_{timestamp}.xlsx"

        return FileResponse(
            path=temp_file_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка при генерации Excel файла: {str(e)}"
        )