CM_TO_POINTS = 28.35  # 1 см ≈ 28.35 точек

import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
from sqlmodel import select
from models.models import DemoRecord, Upload
from fastapi import HTTPException
from fastapi.responses import FileResponse
from tempfile import NamedTemporaryFile


# Разделяем префикс и числа
def split_item(item):
    """
    R1, R2, R3 -> R1-R3
    Недоработка R1, R? -> R1
    """
    if not item:
        return None, None

    prefix = ''
    num_str = ''
    for char in item:
        if char.isalpha():
            prefix += char
        elif char.isdigit():
            num_str += char
        else:
            # Если встретился небуквенный и нецифровой символ (например, '?'), пропускаем
            pass

    # Если нет числовой части, возвращаем None
    num = int(num_str) if num_str else None
    return prefix, num


def shorten_ranges(s):
    if not s:
        return ""

    items = [item.strip() for item in s.split(",")]
    if not items:
        return ""

    # Проверяем, есть ли числовая часть хотя бы у одного элемента
    has_numbers = any(split_item(item)[1] is not None for item in items)

    if not has_numbers:
        # Если ни у одного элемента нет, просто возвращаем исходную строку
        return s

    # Фильтруем элементы, оставляя только те, у которых есть числовая часть
    valid_items = []
    for item in items:
        prefix, num = split_item(item)
        if num is not None:
            valid_items.append((prefix, num))

    if not valid_items:
        return ""

    # Получаем префикс (предполагаем, что он одинаковый для всех элементов)
    prefix = valid_items[0][0]
    numbers = [num for _, num in valid_items]

    ranges = []
    start = numbers[0]
    prev = start

    for num in numbers[1:]:
        if num == prev + 1:
            prev = num
        else:
            if start == prev:
                ranges.append(f"{prefix}{start}")
            else:
                ranges.append(f"{prefix}{start}-{prefix}{prev}")
            start = num
            prev = num

    # Добавляем последний диапазон
    if start == prev:
        ranges.append(f"{prefix}{start}")
    else:
        ranges.append(f"{prefix}{start}-{prefix}{prev}")

    return ", ".join(ranges)


def generate_excel_for_upload(upload_id: int, session):
    """
    Генерирует Excel файл для указанного upload_id и возвращает FileResponse
    """
    try:
        # Получаем записи
        records = session.exec(
            select(DemoRecord)
            .where(DemoRecord.upload_id == upload_id)
        ).all()

        if not records:
            raise HTTPException(status_code=404, detail="Записей нет")

        # Получаем информацию о загрузке
        upload = session.get(Upload, upload_id)
        if not upload:
            raise HTTPException(status_code=404, detail="Загрузка не найдена")

        # Создаем Excel табличку
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM Data"
        default_font = Font(name='GOST Type A', size=12, italic=True)
        align_center = Alignment(horizontal="center", vertical="center")

        # Заголовки
        headers = [
            "Поз. обозначение",
            "Наименование",
            "Кол.",
            "Примечание"
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = default_font
            cell.alignment = align_center  # выравнивание заголовков по центру

        column_widths = {
            'A': 2.0 * CM_TO_POINTS,    # 2.0 см
            'B': 11.0 * CM_TO_POINTS,   # 11.0 см
            'C': 1.0 * CM_TO_POINTS,    # 1.0 см
            'D': 4.5 * CM_TO_POINTS     # 4.5 см
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width / 7

        # Данные
        for record in records:
            row = [
                shorten_ranges(record.designator),
                f"{record.ad_class} {record.ad_bom} {record.ad_ss}".strip(),
                record.quantity,
                record.ad_note
            ]
            ws.append(row)

            current_row = ws.max_row
            for col_idx, cell in enumerate(ws[current_row], start=1):
                cell.font = default_font
                # Столбец A (1) и C (3) — выравнивание по центру
                if col_idx in (1, 3):
                    cell.alignment = align_center

        # Создаем временный файл (кросс-платформенный способ)
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_file_path = temp_file.name
            wb.save(temp_file_path)

        # Генерируем имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"bom_upload_{upload_id}_{timestamp}.xlsx"

        return FileResponse(
            path=temp_file_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка при генерации Excel файла: {str(e)}"
        )
    finally:
        # Удаление временного файла будет выполнено FastAPI автоматически после отправки
        pass


